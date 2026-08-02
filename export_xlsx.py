"""
Exports current results to a single .xlsx workbook (Summary, Model
Comparison, Current Status, Fills, Portfolio History tabs) that opens
directly in Excel/Numbers. Safe to run while main.py is running (SQLite
WAL allows concurrent readers).
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

import storage


def write_sheet(ws, headers: list[str], rows: list[tuple]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)


def portfolio_summary(conn, model: str, mids: dict, starting_cash: float) -> tuple[float, dict, float]:
    row = conn.execute(
        "SELECT cash, positions_json FROM portfolio_snapshots WHERE model=? ORDER BY ts DESC LIMIT 1",
        (model,),
    ).fetchone()
    if not row:
        return starting_cash, {}, starting_cash
    cash, positions_json = row
    positions = json.loads(positions_json)
    mtm = cash + sum(pos * mids.get(t, 0) for t, pos in positions.items())
    return cash, positions, mtm


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Kalshi MM results to .xlsx")
    parser.add_argument("--db-path", default="data/kalshi_mm.db")
    parser.add_argument("--out-path", default="data/exports/kalshi_mm_results.xlsx")
    parser.add_argument("--starting-cash", type=float, default=5000.0)
    args = parser.parse_args()

    conn = storage.connect(args.db_path)

    status_rows = conn.execute(
        """
        SELECT bt.ticker,
               datetime(bt.ts, 'unixepoch', 'localtime') AS book_time,
               bt.best_bid, bt.best_bid_qty, bt.best_ask, bt.best_ask_qty,
               q.bid_price AS our_bid, q.bid_size AS our_bid_size,
               q.ask_price AS our_ask, q.ask_size AS our_ask_size
        FROM book_top bt
        LEFT JOIN quotes q
          ON q.ticker = bt.ticker AND q.ts = (SELECT MAX(ts) FROM quotes WHERE ticker = bt.ticker)
        WHERE bt.ts = (SELECT MAX(ts) FROM book_top WHERE ticker = bt.ticker)
        ORDER BY bt.ticker
        """
    ).fetchall()

    fill_rows = conn.execute(
        "SELECT ticker, datetime(ts, 'unixepoch', 'localtime') AS time, side, price, qty, "
        "cash_after, position_after, model FROM fills ORDER BY ts DESC"
    ).fetchall()

    portfolio_rows = conn.execute(
        "SELECT datetime(ts, 'unixepoch', 'localtime') AS time, cash, positions_json, model "
        "FROM portfolio_snapshots ORDER BY ts DESC"
    ).fetchall()

    mids = {r[0]: (r[2] + r[4]) / 2 / 100 for r in status_rows if r[2] is not None and r[4] is not None}
    cash_c, positions_c, mtm_c = portfolio_summary(conn, "conservative", mids, args.starting_cash)
    cash, positions, mtm = portfolio_summary(conn, "optimistic", mids, args.starting_cash)

    wb = Workbook()

    # Headline: queue-aware (conservative) is the realistic number. Optimistic
    # ignores queue priority (assumes any crossing trade fills us) and is kept
    # only as a reference upper bound — see Model Comparison.
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_sheet(
        ws_summary,
        ["cash", "mark_to_market", "pnl_vs_starting_cash", "starting_cash", "open_positions"],
        [(round(cash_c, 2), round(mtm_c, 2), round(mtm_c - args.starting_cash, 2), args.starting_cash, str(positions_c))],
    )

    ws_compare = wb.create_sheet("Model Comparison")
    write_sheet(
        ws_compare,
        ["model", "cash", "mark_to_market", "pnl_vs_starting_cash", "positions"],
        [
            ("REAL: queue-aware (needs real volume to clear depth ahead of us before we fill)",
             round(cash_c, 2), round(mtm_c, 2), round(mtm_c - args.starting_cash, 2), str(positions_c)),
            ("reference only: optimistic (ignores queue priority, any crossing trade fills us)",
             round(cash, 2), round(mtm, 2), round(mtm - args.starting_cash, 2), str(positions)),
            ("gap (optimistic - queue-aware) — how much the naive model overstates P&L",
             round(cash - cash_c, 2), round(mtm - mtm_c, 2), round((mtm - args.starting_cash) - (mtm_c - args.starting_cash), 2), ""),
        ],
    )

    ws_status = wb.create_sheet("Current Status")
    write_sheet(
        ws_status,
        ["ticker", "book_time", "best_bid", "best_bid_qty", "best_ask", "best_ask_qty",
         "our_bid", "our_bid_size", "our_ask", "our_ask_size"],
        status_rows,
    )

    ws_fills = wb.create_sheet("Fills")
    write_sheet(
        ws_fills,
        ["ticker", "time", "side", "price", "qty", "cash_after", "position_after", "model"],
        fill_rows,
    )

    ws_portfolio = wb.create_sheet("Portfolio History")
    write_sheet(ws_portfolio, ["time", "cash", "positions_json", "model"], portfolio_rows)

    out_path = Path(args.out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    conn.close()
    print(f"wrote {out_path} ({len(status_rows)} markets, {len(fill_rows)} fills, "
          f"gap={mtm - mtm_c:+.2f})")


if __name__ == "__main__":
    main()
